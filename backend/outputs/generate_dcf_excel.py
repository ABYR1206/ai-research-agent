"""专业级 DCF Excel 模型生成器。

对外暴露单一函数 `generate_dcf_excel`：接收两个 dict（historical_data + assumptions）
和元数据，输出 7-sheet 投行风格 .xlsx 文件。

设计原则：
- Forecast 与 DCF sheet 全部用 Excel 公式（修改 Assumptions 后整模型自动重算）
- Sensitivity 用 SUMPRODUCT 展开的公式实现，FCF 改了也会重算
- 三色单元格分区：浅黄=输入、浅蓝=公式、浅绿=输出
- 商务深蓝 (#1F3864) 主色 + 标题白字大字号
"""
from __future__ import annotations
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


# ============================================================================
# 样式常量
# ============================================================================
NAVY = "1F3864"
NAVY_LIGHT = "2E5090"
ACCENT_GOLD = "BF9000"
INPUT_BG = "FFF2CC"     # 浅黄：输入假设
FORMULA_BG = "DDEBF7"   # 浅蓝：公式
OUTPUT_BG = "E2EFDA"    # 浅绿：关键输出
SECTION_BG = "D9E1F2"   # 区块小标题
HEADER_BG = "BDD7EE"    # 年份表头
WHITE = "FFFFFF"
GREY = "F2F2F2"

FMT_MONEY = '#,##0.0;[Red](#,##0.0);"–"'
FMT_PCT = '0.0%;[Red]-0.0%;"–"'
FMT_PRICE = '"$"#,##0.00;[Red]("$"#,##0.00);"–"'
FMT_INT = '#,##0;[Red](#,##0)'
FMT_FACTOR = '0.0000'

_thin = Side(border_style="thin", color="BFBFBF")
_medium = Side(border_style="medium", color=NAVY)
_thick = Side(border_style="thick", color=NAVY)
BORDER_THIN = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
BORDER_MEDIUM = Border(left=_medium, right=_medium, top=_medium, bottom=_medium)
BORDER_TOP = Border(top=_medium)
BORDER_BOTTOM = Border(bottom=_medium)

FONT_TITLE = Font(name="Calibri", size=28, bold=True, color=WHITE)
FONT_SUBTITLE = Font(name="Calibri", size=14, italic=True, color=WHITE)
FONT_BANNER = Font(name="Calibri", size=13, bold=True, color=WHITE)
FONT_SECTION = Font(name="Calibri", size=11, bold=True, color=NAVY)
FONT_HEADER = Font(name="Calibri", size=10, bold=True, color=NAVY)
FONT_LABEL = Font(name="Calibri", size=10, bold=True, color="1F1F1F")
FONT_BODY = Font(name="Calibri", size=10, color="1F1F1F")
FONT_OUTPUT = Font(name="Calibri", size=11, bold=True, color="1F1F1F")
FONT_HERO = Font(name="Calibri", size=20, bold=True, color=NAVY)
FONT_DISCLAIMER = Font(name="Calibri", size=8, italic=True, color="666666")

FILL_NAVY = PatternFill("solid", fgColor=NAVY)
FILL_NAVY_LIGHT = PatternFill("solid", fgColor=NAVY_LIGHT)
FILL_INPUT = PatternFill("solid", fgColor=INPUT_BG)
FILL_FORMULA = PatternFill("solid", fgColor=FORMULA_BG)
FILL_OUTPUT = PatternFill("solid", fgColor=OUTPUT_BG)
FILL_SECTION = PatternFill("solid", fgColor=SECTION_BG)
FILL_HEADER = PatternFill("solid", fgColor=HEADER_BG)
FILL_GREY = PatternFill("solid", fgColor=GREY)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", indent=1)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center", indent=1)


# ============================================================================
# Assumptions sheet 单元格地址（其他 sheet 通过这些常量交叉引用）
# ============================================================================
# Drivers 区块：B4-B11
A_WACC = "Assumptions!$B$4"
A_TG = "Assumptions!$B$5"
A_TAX = "Assumptions!$B$6"
A_DA = "Assumptions!$B$7"
A_CAPEX = "Assumptions!$B$8"
A_NWC = "Assumptions!$B$9"
A_NETDEBT = "Assumptions!$B$10"
A_SHARES = "Assumptions!$B$11"

# Per-year forecast 行：行 16 = Revenue Growth, 行 17 = EBIT Margin
# 列 B..F = Y1..Y5
def a_growth(i: int) -> str:
    return f"Assumptions!${get_column_letter(2 + i)}$16"

def a_margin(i: int) -> str:
    return f"Assumptions!${get_column_letter(2 + i)}$17"


# ============================================================================
# 通用辅助
# ============================================================================
def _set(ws: Worksheet, addr_or_row, col=None, *, value=None, fill=None,
         font=None, fmt=None, border=None, align=None, merge=None):
    """统一的单元格写入助手。"""
    cell = ws[addr_or_row] if col is None else ws.cell(row=addr_or_row, column=col)
    if value is not None:
        cell.value = value
    if fill is not None:
        cell.fill = fill
    if font is not None:
        cell.font = font
    if fmt is not None:
        cell.number_format = fmt
    if border is not None:
        cell.border = border
    if align is not None:
        cell.alignment = align
    if merge is not None:
        ws.merge_cells(merge)
    return cell


def _banner(ws: Worksheet, text: str, last_col: int = 8, height: int = 32):
    """每个 sheet 顶部的深蓝大 banner。"""
    last = get_column_letter(last_col)
    ws.merge_cells(f"A1:{last}1")
    c = ws["A1"]
    c.value = text
    c.fill = FILL_NAVY
    c.font = FONT_BANNER
    c.alignment = ALIGN_CENTER
    ws.row_dimensions[1].height = height


def _section(ws: Worksheet, row: int, text: str, last_col: int = 8):
    """子区块小标题（浅蓝底深蓝字）。"""
    last = get_column_letter(last_col)
    ws.merge_cells(f"A{row}:{last}{row}")
    c = ws.cell(row=row, column=1, value=text)
    c.fill = FILL_SECTION
    c.font = FONT_SECTION
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 22


def _label(ws: Worksheet, row: int, text: str):
    c = ws.cell(row=row, column=1, value=text)
    c.font = FONT_LABEL
    c.alignment = ALIGN_LEFT
    c.border = BORDER_THIN
    return c


def _input(ws: Worksheet, row: int, col: int, value, fmt: str):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = FILL_INPUT
    c.font = FONT_BODY
    c.number_format = fmt
    c.border = BORDER_THIN
    c.alignment = ALIGN_RIGHT
    return c


def _formula(ws: Worksheet, row: int, col: int, formula: str, fmt: str):
    c = ws.cell(row=row, column=col, value=formula)
    c.fill = FILL_FORMULA
    c.font = FONT_BODY
    c.number_format = fmt
    c.border = BORDER_THIN
    c.alignment = ALIGN_RIGHT
    return c


def _output(ws: Worksheet, row: int, col: int, value_or_formula, fmt: str, *, hero=False):
    c = ws.cell(row=row, column=col, value=value_or_formula)
    c.fill = FILL_OUTPUT
    c.number_format = fmt
    c.alignment = ALIGN_RIGHT
    c.border = BORDER_MEDIUM if hero else BORDER_THIN
    c.font = Font(name="Calibri", size=14 if hero else 11, bold=True, color=NAVY if hero else "1F1F1F")
    return c


def _year_header(ws: Worksheet, row: int, col: int, value):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = FILL_HEADER
    c.font = FONT_HEADER
    c.alignment = ALIGN_CENTER
    c.border = BORDER_THIN
    return c


def _box_merged(ws: Worksheet, range_str: str, border: Border) -> None:
    """给合并区域的每个 cell 设置 border，让 box 外框完整显示。

    openpyxl 的 merged cell 边框只在左上角 cell 设是不够的 —— Excel
    渲染时只画左上一格的边，box 就会看起来'格子不全'。必须遍历整个
    范围给每个 cell 都设 border。
    """
    from openpyxl.utils import range_boundaries
    min_col, min_row, max_col, max_row = range_boundaries(range_str)
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            ws.cell(row=r, column=c).border = border


def _autosize(ws: Worksheet, min_w: int = 12, max_w: int = 30):
    """根据每列内容估算合适列宽。

    关键：公式单元格不能跳过 —— 否则全表都是公式的 sheet（DCF/Forecast）
    会按标签宽度算，导致 ######。改为按 number_format 估算输出宽度。
    """
    from openpyxl.cell.cell import MergedCell

    # 每种 number_format 在最坏情况下需要的字符数。
    # 注意：openpyxl 的 width 单位 ≠ 纯字符数；加粗字体在 Excel 中实际
    # 占用更多像素。下列估算已包含负数红字括号和加粗 padding（×1.4 安全系数）。
    NF_MIN_WIDTH = {
        FMT_MONEY: 20,    # "(1,813,239.0)" 加粗后约 20 单位
        FMT_PRICE: 16,    # "$1,234,567.89"
        FMT_PCT: 12,      # "(123.4%)"
        FMT_INT: 16,
        FMT_FACTOR: 11,   # "0.9999" + padding
    }

    widths: dict[str, int] = {}
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c, MergedCell) or c.value is None:
                continue
            s = str(c.value)
            if s.startswith("="):
                est = NF_MIN_WIDTH.get(c.number_format, min_w)
            else:
                est = len(s)
                # 数字类型的字面值也要确保格式后宽度足够
                if c.number_format in NF_MIN_WIDTH:
                    est = max(est, NF_MIN_WIDTH[c.number_format])
            letter = c.column_letter
            if est > widths.get(letter, min_w):
                widths[letter] = est

    for letter, w in widths.items():
        ws.column_dimensions[letter].width = min(w + 2, max_w)


# ============================================================================
# Sheet 1: Cover
# ============================================================================
def _build_cover(wb: Workbook, company_name: str, ticker: str):
    ws = wb.create_sheet("Cover", 0)
    ws.sheet_view.showGridLines = False

    # 顶部巨幅标题 (A1:H6)
    ws.merge_cells("A1:H6")
    t = ws["A1"]
    t.value = f"{company_name}\n({ticker})"
    t.fill = FILL_NAVY
    t.font = FONT_TITLE
    t.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r in range(1, 7):
        ws.row_dimensions[r].height = 32

    ws.merge_cells("A7:H7")
    s = ws["A7"]
    s.value = "Discounted Cash Flow Valuation Model"
    s.fill = FILL_NAVY_LIGHT
    s.font = FONT_SUBTITLE
    s.alignment = ALIGN_CENTER
    ws.row_dimensions[7].height = 26

    # 生成日期 + 模型版本
    ws["A9"] = "Report Date"
    ws["A9"].font = FONT_HEADER
    ws["B9"] = date.today().isoformat()
    ws["B9"].font = FONT_BODY
    ws["G9"] = "Model Version"
    ws["G9"].font = FONT_HEADER
    ws["G9"].alignment = ALIGN_RIGHT
    ws["H9"] = "v1.0"
    ws["H9"].font = FONT_BODY
    ws["H9"].alignment = ALIGN_RIGHT

    # Key Valuation Output 框 (B12:G22)
    _section(ws, 11, "Key Valuation Output", last_col=8)

    box_rows = [
        ("Enterprise Value (EV)", "='DCF Valuation'!B13", FMT_MONEY, False),
        ("Less: Net Debt", "='DCF Valuation'!B14", FMT_MONEY, False),
        ("Equity Value", "='DCF Valuation'!B15", FMT_MONEY, False),
        ("Shares Outstanding (M)", "='DCF Valuation'!B16", FMT_MONEY, False),
        ("Implied Share Price", "='DCF Valuation'!B17", FMT_PRICE, True),
        ("Current Share Price", "='DCF Valuation'!B18", FMT_PRICE, False),
        ("Upside / (Downside)", "='DCF Valuation'!B19", FMT_PCT, True),
    ]
    for i, (label, formula, fmt, hero) in enumerate(box_rows):
        r = 13 + i
        # 左侧 label box: B:D
        ws.merge_cells(f"B{r}:D{r}")
        lc = ws.cell(row=r, column=2, value=label)
        lc.font = FONT_LABEL
        lc.alignment = ALIGN_LEFT
        # 给整段合并范围画 box 边框（hero 行用粗边框统一视觉）
        _box_merged(ws, f"B{r}:D{r}", BORDER_MEDIUM if hero else BORDER_THIN)

        # 右侧数值 box: E:G
        ws.merge_cells(f"E{r}:G{r}")
        vc = ws.cell(row=r, column=5, value=formula)
        vc.number_format = fmt
        vc.alignment = ALIGN_RIGHT
        _box_merged(ws, f"E{r}:G{r}", BORDER_MEDIUM if hero else BORDER_THIN)

        if hero:
            # hero 行：填充绿底 + 大字号
            for c in range(5, 8):
                ws.cell(row=r, column=c).fill = FILL_OUTPUT
            vc.font = Font(name="Calibri", size=16, bold=True, color=NAVY)
            ws.row_dimensions[r].height = 28
        else:
            vc.font = FONT_OUTPUT
            ws.row_dimensions[r].height = 22

    # Methodology Note
    _section(ws, 22, "Methodology", last_col=8)
    notes = [
        "• 5-year explicit forecast period with Free Cash Flow projected from revenue growth, EBIT margin, tax, D&A, capex, and NWC assumptions.",
        "• Terminal value calculated via the Gordon Growth model: TV = FCF₅ × (1 + g) / (WACC − g).",
        "• Enterprise Value = Sum of discounted FCFs + Discounted Terminal Value.",
        "• Equity Value = Enterprise Value − Net Debt (Total Debt − Cash).",
        "• Implied Share Price = Equity Value ÷ Shares Outstanding.",
    ]
    for i, note in enumerate(notes):
        r = 23 + i
        ws.merge_cells(f"A{r}:H{r}")
        c = ws.cell(row=r, column=1, value=note)
        c.font = FONT_BODY
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False, indent=1)

    # Disclaimer
    disc_r = 30
    ws.merge_cells(f"A{disc_r}:H{disc_r}")
    d = ws.cell(row=disc_r, column=1,
                value="Disclaimer: This model is generated by AI Research Agent for academic / demonstration purposes only. Figures are illustrative and do not constitute investment advice.")
    d.font = FONT_DISCLAIMER
    d.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[disc_r].height = 30

    # 列宽
    for col, w in zip("ABCDEFGH", [16, 16, 14, 14, 14, 14, 16, 14]):
        ws.column_dimensions[col].width = w


# ============================================================================
# Sheet 2: Assumptions
# ============================================================================
def _build_assumptions(wb: Workbook, assumptions: dict, historical: dict, current_price: float):
    ws = wb.create_sheet("Assumptions")
    ws.sheet_view.showGridLines = False
    _banner(ws, "Model Assumptions  —  edit yellow cells; the entire model recalculates automatically", last_col=8)

    # Drivers 块
    _section(ws, 3, "Valuation Drivers")
    drivers = [
        ("WACC (Discount Rate)", assumptions["wacc"], FMT_PCT),
        ("Terminal Growth Rate (g)", assumptions["terminal_growth_rate"], FMT_PCT),
        ("Effective Tax Rate", assumptions["tax_rate"], FMT_PCT),
        ("D&A % of Revenue", assumptions["da_percent_revenue"], FMT_PCT),
        ("Capex % of Revenue", assumptions["capex_percent_revenue"], FMT_PCT),
        ("ΔNWC % of Revenue", assumptions["nwc_percent_revenue"], FMT_PCT),
    ]
    for i, (label, val, fmt) in enumerate(drivers):
        r = 4 + i
        _label(ws, r, label)
        _input(ws, r, 2, val, fmt)

    # Balance Sheet Items (net debt 由 historical cash/debt 推导，可手改)
    last_cash = historical["cash"][-1]
    last_debt = historical["total_debt"][-1]
    net_debt = max(0.0, last_debt - last_cash)
    _label(ws, 10, "Net Debt (Total Debt − Cash, M)")
    _input(ws, 10, 2, net_debt, FMT_MONEY)
    _label(ws, 11, "Shares Outstanding (M)")
    _input(ws, 11, 2, historical["shares_outstanding"][-1], FMT_MONEY)
    _label(ws, 12, "Current Share Price")
    _input(ws, 12, 2, current_price, FMT_PRICE)

    # Per-year growth & margin
    _section(ws, 14, "Forecast Period Drivers (Year-by-Year)")
    n = len(assumptions["forecast_years"])
    _year_header(ws, 15, 1, "Year")
    for i, y in enumerate(assumptions["forecast_years"]):
        _year_header(ws, 15, 2 + i, f"{y}E")

    _label(ws, 16, "Revenue Growth")
    for i, g in enumerate(assumptions["revenue_growth"]):
        _input(ws, 16, 2 + i, g, FMT_PCT)
    _label(ws, 17, "EBIT Margin")
    for i, m in enumerate(assumptions["ebit_margin"]):
        _input(ws, 17, 2 + i, m, FMT_PCT)

    # Legend
    _section(ws, 20, "Legend")
    legend = [
        ("Input cell (editable)", FILL_INPUT),
        ("Formula cell (do not edit)", FILL_FORMULA),
        ("Key output cell", FILL_OUTPUT),
    ]
    for i, (text, fill) in enumerate(legend):
        r = 21 + i
        c = ws.cell(row=r, column=1, value=" ")
        c.fill = fill
        c.border = BORDER_THIN
        ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width or 30, 30)
        lc = ws.cell(row=r, column=2, value=text)
        lc.font = FONT_BODY
        lc.alignment = ALIGN_LEFT

    _autosize(ws, min_w=16, max_w=34)
    ws.column_dimensions["A"].width = 36
    ws.freeze_panes = "B4"


# ============================================================================
# Sheet 3: Historical Financials
# ============================================================================
def _build_historical(wb: Workbook, historical: dict):
    ws = wb.create_sheet("Historical Financials")
    ws.sheet_view.showGridLines = False
    n = len(historical["years"])
    _banner(ws, f"Historical Financials  ({n} fiscal years, USD millions)", last_col=1 + n)

    # 表头
    _year_header(ws, 3, 1, "Line Item")
    for i, y in enumerate(historical["years"]):
        _year_header(ws, 3, 2 + i, y)

    items = [
        ("Revenue", "revenue", FMT_MONEY),
        ("EBIT", "ebit", FMT_MONEY),
        ("Net Income", "net_income", FMT_MONEY),
        ("D&A", "depreciation_amortization", FMT_MONEY),
        ("Capex", "capex", FMT_MONEY),
        ("Δ in NWC", "change_in_nwc", FMT_MONEY),
        ("Free Cash Flow", "free_cash_flow", FMT_MONEY),
        ("Cash & Equivalents", "cash", FMT_MONEY),
        ("Total Debt", "total_debt", FMT_MONEY),
        ("Shares Outstanding (M)", "shares_outstanding", FMT_MONEY),
    ]
    for ri, (label, key, fmt) in enumerate(items):
        r = 4 + ri
        _label(ws, r, label)
        for i, v in enumerate(historical[key]):
            c = ws.cell(row=r, column=2 + i, value=v)
            c.fill = FILL_GREY if ri % 2 == 0 else PatternFill()
            c.number_format = fmt
            c.font = FONT_BODY
            c.border = BORDER_THIN
            c.alignment = ALIGN_RIGHT

    # Calculated ratios (公式版)
    ratios_start = 4 + len(items) + 1
    _section(ws, ratios_start, "Calculated Ratios", last_col=1 + n)
    ratio_specs = [
        ("Revenue Growth YoY", lambda i, col: None if i == 0 else f"={col}4/{get_column_letter(2 + i - 1)}4-1", FMT_PCT),
        ("EBIT Margin", lambda i, col: f"={col}5/{col}4", FMT_PCT),
        ("D&A % Revenue", lambda i, col: f"={col}7/{col}4", FMT_PCT),
        ("Capex % Revenue", lambda i, col: f"={col}8/{col}4", FMT_PCT),
        ("FCF Margin", lambda i, col: f"={col}10/{col}4", FMT_PCT),
    ]
    for ri, (label, formula_fn, fmt) in enumerate(ratio_specs):
        r = ratios_start + 1 + ri
        _label(ws, r, label)
        for i in range(n):
            col = get_column_letter(2 + i)
            f = formula_fn(i, col)
            if f is None:
                c = ws.cell(row=r, column=2 + i, value="–")
                c.font = FONT_BODY
                c.alignment = ALIGN_CENTER
                c.border = BORDER_THIN
            else:
                _formula(ws, r, 2 + i, f, fmt)

    _autosize(ws, min_w=15, max_w=28)
    ws.column_dimensions["A"].width = 24
    ws.freeze_panes = "B4"


# ============================================================================
# Sheet 4: Forecast
# ============================================================================
def _build_forecast(wb: Workbook, historical: dict, assumptions: dict):
    """5-year forecast with EVERY projected cell as a real formula."""
    ws = wb.create_sheet("Forecast")
    ws.sheet_view.showGridLines = False

    n = len(assumptions["forecast_years"])
    last_year = historical["years"][-1]
    last_rev = historical["revenue"][-1]

    _banner(ws, "Forecast  —  formulas linked to Assumptions sheet", last_col=2 + n)

    # 表头：A=Item, B=LY actual, C..G = Y1..Y5
    _year_header(ws, 3, 1, "Line Item")
    _year_header(ws, 3, 2, f"{last_year}A")
    for i, y in enumerate(assumptions["forecast_years"]):
        _year_header(ws, 3, 3 + i, f"{y}E")

    # 行布局
    R_GROWTH, R_REV, R_MARGIN, R_EBIT, R_TAX, R_NOPAT, R_DA, R_CAPEX, R_NWC, R_DNWC, R_FCF = \
        4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14

    _label(ws, R_GROWTH, "Revenue Growth %")
    _label(ws, R_REV, "Revenue")
    _label(ws, R_MARGIN, "EBIT Margin %")
    _label(ws, R_EBIT, "EBIT")
    _label(ws, R_TAX, "Tax on EBIT")
    _label(ws, R_NOPAT, "NOPAT")
    _label(ws, R_DA, "D&A")
    _label(ws, R_CAPEX, "Capex")
    _label(ws, R_NWC, "Net Working Capital")
    _label(ws, R_DNWC, "Δ in NWC")
    _label(ws, R_FCF, "Free Cash Flow")

    # LY 实际值（B 列，灰色背景）
    last_ebit = historical["ebit"][-1]
    last_da = historical["depreciation_amortization"][-1]
    last_capex = historical["capex"][-1]
    last_nwc_proxy = last_rev * assumptions["nwc_percent_revenue"]

    actuals = {R_REV: last_rev, R_EBIT: last_ebit, R_DA: last_da,
               R_CAPEX: last_capex, R_NWC: last_nwc_proxy}
    for r, v in actuals.items():
        c = ws.cell(row=r, column=2, value=v)
        c.fill = FILL_GREY
        c.font = FONT_BODY
        c.number_format = FMT_MONEY
        c.border = BORDER_THIN
        c.alignment = ALIGN_RIGHT
    # margin LY shown as ratio (informational)
    c = ws.cell(row=R_MARGIN, column=2, value=f"=IFERROR(B{R_EBIT}/B{R_REV},0)")
    c.fill = FILL_GREY
    c.font = FONT_BODY
    c.number_format = FMT_PCT
    c.border = BORDER_THIN
    c.alignment = ALIGN_RIGHT

    # 预测列 C..G
    for i in range(n):
        col = get_column_letter(3 + i)
        prev = get_column_letter(3 + i - 1) if i > 0 else "B"

        _formula(ws, R_GROWTH, 3 + i, f"={a_growth(i)}", FMT_PCT)
        _formula(ws, R_REV, 3 + i, f"={prev}{R_REV}*(1+{col}{R_GROWTH})", FMT_MONEY)
        _formula(ws, R_MARGIN, 3 + i, f"={a_margin(i)}", FMT_PCT)
        _formula(ws, R_EBIT, 3 + i, f"={col}{R_REV}*{col}{R_MARGIN}", FMT_MONEY)
        _formula(ws, R_TAX, 3 + i, f"={col}{R_EBIT}*{A_TAX}", FMT_MONEY)
        _formula(ws, R_NOPAT, 3 + i, f"={col}{R_EBIT}-{col}{R_TAX}", FMT_MONEY)
        _formula(ws, R_DA, 3 + i, f"={col}{R_REV}*{A_DA}", FMT_MONEY)
        _formula(ws, R_CAPEX, 3 + i, f"={col}{R_REV}*{A_CAPEX}", FMT_MONEY)
        _formula(ws, R_NWC, 3 + i, f"={col}{R_REV}*{A_NWC}", FMT_MONEY)
        _formula(ws, R_DNWC, 3 + i, f"={col}{R_NWC}-{prev}{R_NWC}", FMT_MONEY)
        # FCF = NOPAT + D&A - Capex - ΔNWC，highlight 输出
        _output(ws, R_FCF, 3 + i,
                f"={col}{R_NOPAT}+{col}{R_DA}-{col}{R_CAPEX}-{col}{R_DNWC}",
                FMT_MONEY)

    # 分隔线
    for col_idx in range(1, 3 + n):
        ws.cell(row=R_FCF + 1, column=col_idx).border = BORDER_TOP

    _autosize(ws, min_w=16, max_w=28)
    ws.column_dimensions["A"].width = 26
    ws.freeze_panes = "C4"


# ============================================================================
# Sheet 5: DCF Valuation
# ============================================================================
def _build_dcf(wb: Workbook, assumptions: dict):
    ws = wb.create_sheet("DCF Valuation")
    ws.sheet_view.showGridLines = False

    n = len(assumptions["forecast_years"])
    _banner(ws, "DCF Valuation  —  Implied Share Price (all calculations are live formulas)", last_col=1 + n)

    # 表头
    _year_header(ws, 3, 1, "Item")
    for i, y in enumerate(assumptions["forecast_years"]):
        _year_header(ws, 3, 2 + i, f"{y}E")

    _label(ws, 4, "Forecast Year (t)")
    _label(ws, 5, "Free Cash Flow")
    _label(ws, 6, "Discount Factor")
    _label(ws, 7, "PV of FCF")

    for i in range(n):
        col = get_column_letter(2 + i)
        fc_col = get_column_letter(3 + i)  # Forecast sheet 中 Y_i 列
        # t
        c = ws.cell(row=4, column=2 + i, value=i + 1)
        c.alignment = ALIGN_CENTER
        c.font = FONT_BODY
        c.border = BORDER_THIN
        # FCF reference
        _formula(ws, 5, 2 + i, f"=Forecast!{fc_col}14", FMT_MONEY)
        # Discount Factor
        _formula(ws, 6, 2 + i, f"=1/(1+{A_WACC})^{col}4", FMT_FACTOR)
        # PV of FCF
        _formula(ws, 7, 2 + i, f"={col}5*{col}6", FMT_MONEY)

    last_col = get_column_letter(1 + n)

    # 汇总块（B 列）
    _section(ws, 9, "Valuation Summary", last_col=1 + n)
    _label(ws, 10, "Σ PV of Explicit-Period FCF")
    _formula(ws, 10, 2, f"=SUM(B7:{last_col}7)", FMT_MONEY)
    _label(ws, 11, "Terminal Value (Gordon Growth)")
    _formula(ws, 11, 2,
             f"={last_col}5*(1+{A_TG})/({A_WACC}-{A_TG})",
             FMT_MONEY)
    _label(ws, 12, "PV of Terminal Value")
    _formula(ws, 12, 2, f"=B11/(1+{A_WACC})^{n}", FMT_MONEY)

    # 关键输出
    _label(ws, 13, "Enterprise Value")
    _output(ws, 13, 2, "=B10+B12", FMT_MONEY, hero=True)
    _label(ws, 14, "Less: Net Debt")
    _formula(ws, 14, 2, f"={A_NETDEBT}", FMT_MONEY)
    _label(ws, 15, "Equity Value")
    _output(ws, 15, 2, "=B13-B14", FMT_MONEY, hero=True)
    _label(ws, 16, "Shares Outstanding (M)")
    _formula(ws, 16, 2, f"={A_SHARES}", FMT_MONEY)
    _label(ws, 17, "Implied Share Price")
    _output(ws, 17, 2, "=B15/B16", FMT_PRICE, hero=True)
    _label(ws, 18, "Current Share Price")
    _formula(ws, 18, 2, "=Assumptions!$B$12", FMT_PRICE)
    _label(ws, 19, "Upside / (Downside)")
    _output(ws, 19, 2, "=IFERROR(B17/B18-1,0)", FMT_PCT, hero=True)

    # 高度调整
    for r in (13, 15, 17, 19):
        ws.row_dimensions[r].height = 26

    _autosize(ws, min_w=15, max_w=28)
    ws.column_dimensions["A"].width = 32
    ws.freeze_panes = "B4"


# ============================================================================
# Sheet 6: Sensitivity Analysis（公式版，会随 Forecast/Assumptions 重算）
# ============================================================================
def _build_sensitivity(wb: Workbook, assumptions: dict):
    ws = wb.create_sheet("Sensitivity Analysis")
    ws.sheet_view.showGridLines = False
    _banner(ws, "Sensitivity Analysis  —  Implied Share Price by WACC × Terminal Growth", last_col=9)

    # 网格设定：WACC 7%→13%（步长 1%），g 1%→4%（步长 0.5%）
    wacc_axis = [round(0.07 + i * 0.01, 4) for i in range(7)]
    growth_axis = [round(0.01 + i * 0.005, 4) for i in range(7)]
    n_fcf = len(assumptions["forecast_years"])

    # 表头说明
    ws.merge_cells("A3:I3")
    c = ws.cell(row=3, column=1,
                value="Each cell = full DCF recomputed with the row's WACC and the column's terminal growth rate; live formulas link to the Forecast sheet's FCF row.")
    c.font = FONT_DISCLAIMER
    c.alignment = ALIGN_CENTER

    # 左上角
    corner = ws.cell(row=5, column=1, value="WACC  \\  g")
    corner.fill = FILL_NAVY
    corner.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
    corner.alignment = ALIGN_CENTER
    corner.border = BORDER_THIN

    # 列头：g
    for j, g in enumerate(growth_axis):
        _year_header(ws, 5, 2 + j, g)
        ws.cell(row=5, column=2 + j).number_format = FMT_PCT
    # 行头：WACC
    for i, w in enumerate(wacc_axis):
        _year_header(ws, 6 + i, 1, w)
        ws.cell(row=6 + i, column=1).number_format = FMT_PCT

    # 单元格公式：对每个 (W, g)，重新折现 Forecast 的 FCF + 终值
    # PV_FCF = Σ FCF_t / (1+W)^t   （展开 5 项）
    # TV = FCF_5 * (1+g)/(W-g)
    # PV_TV = TV / (1+W)^5
    # Implied = (PV_FCF + PV_TV - NetDebt) / Shares
    #
    # FCF 在 Forecast!C14..G14（n=5 时）
    fcf_cells = [f"Forecast!{get_column_letter(3 + k)}$14" for k in range(n_fcf)]
    for i in range(len(wacc_axis)):
        w_ref = f"$A{6 + i}"
        for j in range(len(growth_axis)):
            g_ref = f"{get_column_letter(2 + j)}$5"
            terms = [f"{fcf}/(1+{w_ref})^{k+1}" for k, fcf in enumerate(fcf_cells)]
            pv_fcf = "+".join(terms)
            tv = f"{fcf_cells[-1]}*(1+{g_ref})/({w_ref}-{g_ref})"
            pv_tv = f"({tv})/(1+{w_ref})^{n_fcf}"
            formula = f"=IFERROR((({pv_fcf})+{pv_tv}-{A_NETDEBT})/{A_SHARES},0)"
            c = ws.cell(row=6 + i, column=2 + j, value=formula)
            c.number_format = FMT_PRICE
            c.font = FONT_BODY
            c.border = BORDER_THIN
            c.alignment = ALIGN_RIGHT

    # 条件格式色阶
    first_row, last_row = 6, 6 + len(wacc_axis) - 1
    rng = f"B{first_row}:H{last_row}"
    ws.conditional_formatting.add(rng, ColorScaleRule(
        start_type="min", start_color="F8696B",
        mid_type="percentile", mid_value=50, mid_color="FFEB84",
        end_type="max", end_color="63BE7B",
    ))

    # 提示当前 Assumptions 中心点
    ws.cell(row=last_row + 2, column=1,
            value=f"Current model assumes WACC = {assumptions['wacc']*100:.1f}%  ·  Terminal Growth = {assumptions['terminal_growth_rate']*100:.1f}%  →  see DCF Valuation sheet for base-case Implied Share Price."
            ).font = FONT_DISCLAIMER
    ws.merge_cells(start_row=last_row + 2, start_column=1, end_row=last_row + 2, end_column=9)

    # 列宽
    ws.column_dimensions["A"].width = 14
    for col_letter in "BCDEFGH":
        ws.column_dimensions[col_letter].width = 14

    ws.freeze_panes = "B6"


# ============================================================================
# Sheet 7: Charts
# ============================================================================
def _build_charts(wb: Workbook, historical: dict, assumptions: dict):
    ws = wb.create_sheet("Charts")
    ws.sheet_view.showGridLines = False
    _banner(ws, "Charts  —  Revenue & Free Cash Flow Trends", last_col=10)

    # 数据区 A3:D? — Year / Type / Revenue / FCF
    _year_header(ws, 3, 1, "Year")
    _year_header(ws, 3, 2, "Period")
    _year_header(ws, 3, 3, "Revenue")
    _year_header(ws, 3, 4, "Free Cash Flow")

    r = 4
    for i, y in enumerate(historical["years"]):
        ws.cell(row=r, column=1, value=y).font = FONT_BODY
        ws.cell(row=r, column=2, value="Actual").font = FONT_BODY
        c = ws.cell(row=r, column=3, value=historical["revenue"][i])
        c.number_format = FMT_MONEY; c.font = FONT_BODY
        c = ws.cell(row=r, column=4, value=historical["free_cash_flow"][i])
        c.number_format = FMT_MONEY; c.font = FONT_BODY
        for col in range(1, 5):
            ws.cell(row=r, column=col).border = BORDER_THIN
        r += 1
    # 预测部分：用公式拉 Forecast sheet（这样改假设后图会重新画）
    for i, y in enumerate(assumptions["forecast_years"]):
        fc_col = get_column_letter(3 + i)
        ws.cell(row=r, column=1, value=y).font = FONT_BODY
        ws.cell(row=r, column=2, value="Forecast").font = FONT_BODY
        c = ws.cell(row=r, column=3, value=f"=Forecast!{fc_col}5")
        c.number_format = FMT_MONEY; c.font = FONT_BODY
        c.fill = FILL_FORMULA
        c = ws.cell(row=r, column=4, value=f"=Forecast!{fc_col}14")
        c.number_format = FMT_MONEY; c.font = FONT_BODY
        c.fill = FILL_FORMULA
        for col in range(1, 5):
            ws.cell(row=r, column=col).border = BORDER_THIN
        r += 1
    end_row = r - 1

    # Revenue bar chart
    bar = BarChart()
    bar.type = "col"
    bar.style = 11
    bar.title = "Revenue Trend (Historical + Forecast)"
    bar.y_axis.title = "USD millions"
    bar.x_axis.title = "Fiscal Year"
    bar.height = 10
    bar.width = 22
    data = Reference(ws, min_col=3, min_row=3, max_col=3, max_row=end_row)
    cats = Reference(ws, min_col=1, min_row=4, max_row=end_row)
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(cats)
    bar.dLbls = DataLabelList(showVal=False)
    ws.add_chart(bar, "F3")

    # FCF line chart
    line = LineChart()
    line.style = 12
    line.title = "Free Cash Flow Trend"
    line.y_axis.title = "USD millions"
    line.x_axis.title = "Fiscal Year"
    line.height = 10
    line.width = 22
    data = Reference(ws, min_col=4, min_row=3, max_col=4, max_row=end_row)
    line.add_data(data, titles_from_data=True)
    line.set_categories(cats)
    s = line.series[0]
    s.smooth = True
    ws.add_chart(line, "F25")

    _autosize(ws, min_w=12, max_w=18)


# ============================================================================
# 对外 API
# ============================================================================
def generate_dcf_excel(
    company_name: str,
    ticker: str,
    historical_data: dict[str, Any],
    assumptions: dict[str, Any],
    output_path: str | Path,
    *,
    current_price: float = 0.0,
) -> Path:
    """生成 7-sheet 投行风格 DCF Excel 模型并保存到磁盘。

    Args:
        company_name: 公司名称（用于 Cover 标题），例如 "Apple Inc."
        ticker: 股票代码，例如 "AAPL"
        historical_data: 历史数据 dict，必须包含以下键：
            years, revenue, ebit, net_income, depreciation_amortization,
            capex, change_in_nwc, free_cash_flow, cash, total_debt,
            shares_outstanding
            （前 5 项为长度 N 的 list，shares_outstanding 也是 list[N]）
        assumptions: 假设 dict，必须包含：
            forecast_years (list[int])
            revenue_growth (list[float], 长度同 forecast_years)
            ebit_margin (list[float], 同上)
            tax_rate (float)
            da_percent_revenue (float)
            capex_percent_revenue (float)
            nwc_percent_revenue (float)
            wacc (float)
            terminal_growth_rate (float)
        output_path: 输出 .xlsx 路径
        current_price: 当前股价（用于上行空间计算），可选

    Returns:
        实际写入磁盘的 Path 对象
    """
    _validate(historical_data, assumptions)
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)  # 删除默认 sheet

    _build_cover(wb, company_name, ticker)
    _build_assumptions(wb, assumptions, historical_data, current_price)
    _build_historical(wb, historical_data)
    _build_forecast(wb, historical_data, assumptions)
    _build_dcf(wb, assumptions)
    _build_sensitivity(wb, assumptions)
    _build_charts(wb, historical_data, assumptions)

    wb.active = 0  # 打开后落在 Cover
    wb.save(path)
    return path


def _validate(historical: dict, assumptions: dict) -> None:
    required_hist = {"years", "revenue", "ebit", "net_income",
                     "depreciation_amortization", "capex", "change_in_nwc",
                     "free_cash_flow", "cash", "total_debt", "shares_outstanding"}
    missing = required_hist - historical.keys()
    if missing:
        raise ValueError(f"historical_data missing keys: {sorted(missing)}")
    n = len(historical["years"])
    for k in required_hist:
        if len(historical[k]) != n:
            raise ValueError(f"historical_data['{k}'] length {len(historical[k])} != years length {n}")

    required_asm = {"forecast_years", "revenue_growth", "ebit_margin", "tax_rate",
                    "da_percent_revenue", "capex_percent_revenue", "nwc_percent_revenue",
                    "wacc", "terminal_growth_rate"}
    missing = required_asm - assumptions.keys()
    if missing:
        raise ValueError(f"assumptions missing keys: {sorted(missing)}")
    nf = len(assumptions["forecast_years"])
    for k in ("revenue_growth", "ebit_margin"):
        if len(assumptions[k]) != nf:
            raise ValueError(f"assumptions['{k}'] length {len(assumptions[k])} != forecast_years length {nf}")
    if assumptions["wacc"] <= assumptions["terminal_growth_rate"]:
        raise ValueError("wacc must be greater than terminal_growth_rate (Gordon model invalid)")
