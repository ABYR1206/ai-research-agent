"""Excel 样式常量：颜色、字体、边框、数字格式。"""
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle

NAVY = "1F3864"
LIGHT_YELLOW = "FFF2CC"   # 输入假设
LIGHT_BLUE = "DDEBF7"     # 公式
LIGHT_GREEN = "E2EFDA"    # 输出结果
WHITE = "FFFFFF"
GREY = "D9D9D9"

FMT_MONEY = '#,##0.0;[Red]-#,##0.0'
FMT_MONEY_M = '#,##0.0," M";[Red]-#,##0.0," M"'
FMT_PCT = '0.0%;[Red]-0.0%'
FMT_RATIO = '0.00"x"'
FMT_PRICE = '"$"#,##0.00;[Red]-"$"#,##0.00'

thin = Side(border_style="thin", color="BFBFBF")
medium = Side(border_style="medium", color=NAVY)
BORDER_ALL = Border(left=thin, right=thin, top=thin, bottom=thin)
BORDER_KEY = Border(left=medium, right=medium, top=medium, bottom=medium)

FONT_TITLE = Font(name="Calibri", size=18, bold=True, color=WHITE)
FONT_H1 = Font(name="Calibri", size=14, bold=True, color=WHITE)
FONT_H2 = Font(name="Calibri", size=11, bold=True, color=NAVY)
FONT_BODY = Font(name="Calibri", size=10)
FONT_BOLD = Font(name="Calibri", size=10, bold=True)
FONT_KEY = Font(name="Calibri", size=14, bold=True, color=NAVY)

FILL_NAVY = PatternFill("solid", fgColor=NAVY)
FILL_YELLOW = PatternFill("solid", fgColor=LIGHT_YELLOW)
FILL_BLUE = PatternFill("solid", fgColor=LIGHT_BLUE)
FILL_GREEN = PatternFill("solid", fgColor=LIGHT_GREEN)
FILL_GREY = PatternFill("solid", fgColor=GREY)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", indent=1)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")


def apply_input(cell):
    cell.fill = FILL_YELLOW; cell.border = BORDER_ALL; cell.font = FONT_BODY
    cell.alignment = ALIGN_RIGHT


def apply_formula(cell):
    cell.fill = FILL_BLUE; cell.border = BORDER_ALL; cell.font = FONT_BODY
    cell.alignment = ALIGN_RIGHT


def apply_output(cell):
    cell.fill = FILL_GREEN; cell.border = BORDER_ALL; cell.font = FONT_BOLD
    cell.alignment = ALIGN_RIGHT


def apply_label(cell):
    cell.font = FONT_BOLD; cell.alignment = ALIGN_LEFT
    cell.border = BORDER_ALL


def apply_header(cell):
    cell.fill = FILL_NAVY; cell.font = FONT_H1
    cell.alignment = ALIGN_CENTER; cell.border = BORDER_ALL


def autosize(ws, min_width=12, max_width=28):
    from openpyxl.utils import get_column_letter
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        max_len = min_width
        for c in col:
            if c.value is not None:
                ln = len(str(c.value))
                if ln > max_len:
                    max_len = ln
        ws.column_dimensions[letter].width = min(max_len + 2, max_width)
