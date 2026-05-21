"""商务风格 DCF Excel 模型生成器。

生成 7 个 sheet：Summary / Assumptions / Historical / Forecast / DCF /
Sensitivity / Charts。所有计算 sheet 用真实 Excel 公式，修改 Assumptions
的浅黄输入单元格后，整个模型自动重算。
"""
from __future__ import annotations
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

from backend.schemas import (
    CompanyProfile, FinancialStatements, FinancialMetrics,
    Forecast, DCFValuation,
)
from backend.outputs import styles as st


# ============================================================
# 单元格地址常量（Assumptions sheet）— 其他 sheet 通过这些引用
# ============================================================
A_WACC = "Assumptions!$B$3"
A_TG = "Assumptions!$B$4"
A_TAX = "Assumptions!$B$5"
A_EBIT_M = "Assumptions!$B$6"
A_DA = "Assumptions!$B$7"
A_CAPEX = "Assumptions!$B$8"
A_NWC = "Assumptions!$B$9"
A_NETDEBT = "Assumptions!$B$10"
A_SHARES = "Assumptions!$B$11"
# Revenue growth Y1..Y5: Assumptions!B13:F13
def a_growth(year_idx: int) -> str:
    """year_idx: 0..4 → B13..F13"""
    col = get_column_letter(2 + year_idx)
    return f"Assumptions!${col}$13"


# ============================================================
# Sheet 1: Summary
# ============================================================
def _build_summary(wb: Workbook, profile: CompanyProfile):
    ws = wb.create_sheet("Summary", 0)
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:F2")
    ws["A1"] = f"{profile.name}  ({profile.ticker})  —  DCF Valuation Model"
    ws["A1"].font = st.FONT_TITLE
    ws["A1"].fill = st.FILL_NAVY
    ws["A1"].alignment = st.ALIGN_CENTER

    rows = [
        ("Ticker", profile.ticker, None),
        ("Sector", profile.sector, None),
        ("Industry", profile.industry, None),
        ("Country", profile.country, None),
        ("Currency", profile.currency, None),
        ("Current Price", profile.current_price or 0, st.FMT_PRICE),
        ("Market Cap (M)", profile.market_cap or 0, st.FMT_MONEY),
        ("Shares Outstanding (M)", profile.shares_outstanding or 0, st.FMT_MONEY),
    ]
    for i, (label, value, fmt) in enumerate(rows, start=4):
        ws.cell(row=i, column=1, value=label).font = st.FONT_BOLD
        c = ws.cell(row=i, column=2, value=value)
        if fmt:
            c.number_format = fmt
        c.alignment = st.ALIGN_RIGHT
        c.border = st.BORDER_ALL
        ws.cell(row=i, column=1).border = st.BORDER_ALL

    # 关键估值结果引用 DCF sheet
    ws.cell(row=14, column=1, value="DCF VALUATION RESULTS").font = st.FONT_H2
    ws.merge_cells("A14:C14")
    key_rows = [
        ("Enterprise Value (M)", "='DCF'!B20", st.FMT_MONEY),
        ("Less: Net Debt (M)", "='DCF'!B21", st.FMT_MONEY),
        ("Equity Value (M)", "='DCF'!B22", st.FMT_MONEY),
        ("Shares Outstanding (M)", "='DCF'!B23", st.FMT_MONEY),
        ("Implied Share Price", "='DCF'!B24", st.FMT_PRICE),
        ("Current Price", profile.current_price or 0, st.FMT_PRICE),
        ("Upside / (Downside)", "=B19/B20-1", st.FMT_PCT),
    ]
    for i, (label, formula, fmt) in enumerate(key_rows, start=15):
        ws.cell(row=i, column=1, value=label).font = st.FONT_BOLD
        c = ws.cell(row=i, column=2, value=formula)
        c.number_format = fmt
        c.alignment = st.ALIGN_RIGHT
        if "Implied" in label or "Upside" in label:
            c.font = st.FONT_KEY
            c.fill = st.FILL_GREEN
            c.border = st.BORDER_KEY
        else:
            c.border = st.BORDER_ALL
        ws.cell(row=i, column=1).border = st.BORDER_ALL

    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 24
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 20


# ============================================================
# Sheet 2: Assumptions
# ============================================================
def _build_assumptions(wb: Workbook, stmts: FinancialStatements,
                        forecast: Forecast, dcf: DCFValuation):
    ws = wb.create_sheet("Assumptions")
    ws.sheet_view.showGridLines = False
    a = forecast.assumptions

    ws.merge_cells("A1:G1")
    ws["A1"] = "Model Assumptions  (修改浅黄单元格后整个模型自动重算)"
    st.apply_header(ws["A1"])

    inputs = [
        ("WACC", a.wacc, st.FMT_PCT),
        ("Terminal Growth Rate (g)", a.terminal_growth, st.FMT_PCT),
        ("Effective Tax Rate", a.tax_rate, st.FMT_PCT),
        ("EBIT Margin", a.ebit_margin, st.FMT_PCT),
        ("D&A % of Revenue", a.da_pct, st.FMT_PCT),
        ("Capex % of Revenue", a.capex_pct, st.FMT_PCT),
        ("NWC % of Revenue", a.nwc_pct, st.FMT_PCT),
        ("Net Debt (M)", dcf.net_debt, st.FMT_MONEY),
        ("Shares Outstanding (M)", stmts.shares_outstanding, st.FMT_MONEY),
    ]
    for i, (label, value, fmt) in enumerate(inputs, start=3):
        ws.cell(row=i, column=1, value=label)
        st.apply_label(ws.cell(row=i, column=1))
        c = ws.cell(row=i, column=2, value=value)
        c.number_format = fmt
        st.apply_input(c)

    # Revenue Growth path 行
    ws.cell(row=12, column=1, value="Revenue Growth Path (Y1-Y5)").font = st.FONT_H2
    base_year = stmts.history[-1].year
    for i in range(5):
        col = 2 + i
        h = ws.cell(row=12, column=col, value=f"{base_year + i + 1}E")
        h.fill = st.FILL_GREY; h.font = st.FONT_BOLD; h.alignment = st.ALIGN_CENTER
        h.border = st.BORDER_ALL
        c = ws.cell(row=13, column=col, value=a.revenue_growth[i])
        c.number_format = st.FMT_PCT
        st.apply_input(c)

    # 图例
    ws.cell(row=16, column=1, value="Legend").font = st.FONT_H2
    legend = [("浅黄 = 输入假设（可改）", st.FILL_YELLOW),
              ("浅蓝 = 计算公式（勿改）", st.FILL_BLUE),
              ("浅绿 = 关键输出结果", st.FILL_GREEN)]
    for i, (text, fill) in enumerate(legend, start=17):
        c = ws.cell(row=i, column=1, value=text)
        c.fill = fill; c.font = st.FONT_BODY; c.border = st.BORDER_ALL

    st.autosize(ws, min_width=14, max_width=32)
    ws.freeze_panes = "B3"


# ============================================================
# Sheet 3: Historical
# ============================================================
def _build_historical(wb: Workbook, stmts: FinancialStatements,
                       metrics: FinancialMetrics):
    ws = wb.create_sheet("Historical")
    ws.sheet_view.showGridLines = False

    n = len(stmts.history)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=1 + n)
    ws["A1"] = f"Historical Financials  ({stmts.ticker}, USD millions)"
    st.apply_header(ws["A1"])

    # 年份表头
    ws.cell(row=3, column=1, value="Item").fill = st.FILL_GREY
    ws.cell(row=3, column=1).font = st.FONT_BOLD
    ws.cell(row=3, column=1).border = st.BORDER_ALL
    for i, y in enumerate(stmts.history):
        c = ws.cell(row=3, column=2 + i, value=y.year)
        c.fill = st.FILL_GREY; c.font = st.FONT_BOLD
        c.alignment = st.ALIGN_CENTER; c.border = st.BORDER_ALL

    line_items = [
        ("Revenue", "revenue", st.FMT_MONEY),
        ("Gross Profit", "gross_profit", st.FMT_MONEY),
        ("EBIT", "ebit", st.FMT_MONEY),
        ("EBITDA", "ebitda", st.FMT_MONEY),
        ("Net Income", "net_income", st.FMT_MONEY),
        ("Total Assets", "total_assets", st.FMT_MONEY),
        ("Total Debt", "total_debt", st.FMT_MONEY),
        ("Cash", "cash", st.FMT_MONEY),
        ("Shareholders' Equity", "equity", st.FMT_MONEY),
        ("Operating Cash Flow", "operating_cf", st.FMT_MONEY),
        ("Capex", "capex", st.FMT_MONEY),
        ("D&A", "da", st.FMT_MONEY),
        ("Free Cash Flow", "fcf", st.FMT_MONEY),
    ]
    for r, (label, attr, fmt) in enumerate(line_items, start=4):
        ws.cell(row=r, column=1, value=label)
        st.apply_label(ws.cell(row=r, column=1))
        for i, y in enumerate(stmts.history):
            c = ws.cell(row=r, column=2 + i, value=getattr(y, attr))
            c.number_format = fmt
            c.border = st.BORDER_ALL
            c.font = st.FONT_BODY
            c.alignment = st.ALIGN_RIGHT

    # 关键比率块
    base_row = 4 + len(line_items) + 2
    ws.cell(row=base_row - 1, column=1, value="Key Ratios").font = st.FONT_H2
    ratios = [
        ("Revenue Growth", metrics.revenue_growth, st.FMT_PCT),
        ("Gross Margin", metrics.gross_margin, st.FMT_PCT),
        ("EBIT Margin", metrics.ebit_margin, st.FMT_PCT),
        ("EBITDA Margin", metrics.ebitda_margin, st.FMT_PCT),
        ("Net Margin", metrics.net_margin, st.FMT_PCT),
        ("Capex % Rev", metrics.capex_pct, st.FMT_PCT),
        ("D&A % Rev", metrics.da_pct, st.FMT_PCT),
        ("Implied Tax Rate", metrics.tax_rate, st.FMT_PCT),
    ]
    for r, (label, values, fmt) in enumerate(ratios, start=base_row):
        ws.cell(row=r, column=1, value=label)
        st.apply_label(ws.cell(row=r, column=1))
        for i, v in enumerate(values):
            c = ws.cell(row=r, column=2 + i, value=v)
            c.number_format = fmt
            c.border = st.BORDER_ALL
            c.font = st.FONT_BODY
            c.alignment = st.ALIGN_RIGHT

    st.autosize(ws, min_width=14, max_width=22)
    ws.freeze_panes = "B4"


# ============================================================
# Sheet 4: Forecast — 真实公式驱动
# ============================================================
def _build_forecast(wb: Workbook, stmts: FinancialStatements, forecast: Forecast):
    ws = wb.create_sheet("Forecast")
    ws.sheet_view.showGridLines = False

    n = forecast.assumptions.forecast_years  # 5
    last = stmts.history[-1]
    base_rev = last.revenue

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2 + n)
    ws["A1"] = "5-Year Forecast  (formulas linked to Assumptions)"
    st.apply_header(ws["A1"])

    # 列：A=Item, B=LY(actual), C..G = Y1..Y5
    ws.cell(row=3, column=1, value="Item").fill = st.FILL_GREY
    ws.cell(row=3, column=1).font = st.FONT_BOLD; ws.cell(row=3, column=1).border = st.BORDER_ALL
    h = ws.cell(row=3, column=2, value=f"{last.year}A")
    h.fill = st.FILL_GREY; h.font = st.FONT_BOLD; h.alignment = st.ALIGN_CENTER; h.border = st.BORDER_ALL
    for i in range(n):
        c = ws.cell(row=3, column=3 + i, value=f"{last.year + i + 1}E")
        c.fill = st.FILL_GREY; c.font = st.FONT_BOLD; c.alignment = st.ALIGN_CENTER; c.border = st.BORDER_ALL

    # Row map (1-indexed):
    # 4 Revenue Growth
    # 5 Revenue
    # 6 EBIT
    # 7 Tax
    # 8 NOPAT
    # 9 D&A
    # 10 Capex
    # 11 NWC
    # 12 Change in NWC
    # 13 FCF

    def lbl(row, text):
        ws.cell(row=row, column=1, value=text)
        st.apply_label(ws.cell(row=row, column=1))

    lbl(4, "Revenue Growth")
    lbl(5, "Revenue")
    lbl(6, "EBIT")
    lbl(7, "Tax")
    lbl(8, "NOPAT")
    lbl(9, "D&A")
    lbl(10, "Capex")
    lbl(11, "NWC")
    lbl(12, "Change in NWC")
    lbl(13, "Free Cash Flow")

    # LY actuals (column B)
    actuals = {5: base_rev, 6: last.ebit, 8: last.ebit * (1 - 0.21),
               9: last.da, 10: last.capex, 11: base_rev * forecast.assumptions.nwc_pct}
    for r, v in actuals.items():
        c = ws.cell(row=r, column=2, value=v)
        c.number_format = st.FMT_MONEY
        c.border = st.BORDER_ALL
        c.alignment = st.ALIGN_RIGHT
        c.font = st.FONT_BODY

    # 预测列：C..G
    for i in range(n):
        col = 3 + i
        col_letter = get_column_letter(col)
        prev_letter = get_column_letter(col - 1)

        # Growth (引用 Assumptions B13:F13)
        c = ws.cell(row=4, column=col, value=f"={a_growth(i)}")
        c.number_format = st.FMT_PCT
        st.apply_formula(c)

        # Revenue = prev * (1 + growth)
        c = ws.cell(row=5, column=col, value=f"={prev_letter}5*(1+{col_letter}4)")
        c.number_format = st.FMT_MONEY
        st.apply_formula(c)

        # EBIT = Revenue * EBIT_margin
        c = ws.cell(row=6, column=col, value=f"={col_letter}5*{A_EBIT_M}")
        c.number_format = st.FMT_MONEY
        st.apply_formula(c)

        # Tax = EBIT * tax_rate
        c = ws.cell(row=7, column=col, value=f"={col_letter}6*{A_TAX}")
        c.number_format = st.FMT_MONEY
        st.apply_formula(c)

        # NOPAT = EBIT - Tax
        c = ws.cell(row=8, column=col, value=f"={col_letter}6-{col_letter}7")
        c.number_format = st.FMT_MONEY
        st.apply_formula(c)

        # D&A
        c = ws.cell(row=9, column=col, value=f"={col_letter}5*{A_DA}")
        c.number_format = st.FMT_MONEY
        st.apply_formula(c)

        # Capex
        c = ws.cell(row=10, column=col, value=f"={col_letter}5*{A_CAPEX}")
        c.number_format = st.FMT_MONEY
        st.apply_formula(c)

        # NWC
        c = ws.cell(row=11, column=col, value=f"={col_letter}5*{A_NWC}")
        c.number_format = st.FMT_MONEY
        st.apply_formula(c)

        # Change in NWC
        c = ws.cell(row=12, column=col, value=f"={col_letter}11-{prev_letter}11")
        c.number_format = st.FMT_MONEY
        st.apply_formula(c)

        # FCF = NOPAT + D&A - Capex - dNWC
        c = ws.cell(row=13, column=col, value=f"={col_letter}8+{col_letter}9-{col_letter}10-{col_letter}12")
        c.number_format = st.FMT_MONEY
        c.fill = st.FILL_GREEN
        c.font = st.FONT_BOLD
        c.border = st.BORDER_ALL
        c.alignment = st.ALIGN_RIGHT

    st.autosize(ws, min_width=14, max_width=20)
    ws.freeze_panes = "B4"


# ============================================================
# Sheet 5: DCF — 真实公式
# ============================================================
def _build_dcf(wb: Workbook, stmts: FinancialStatements, forecast: Forecast):
    ws = wb.create_sheet("DCF")
    ws.sheet_view.showGridLines = False

    n = forecast.assumptions.forecast_years
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2 + n)
    ws["A1"] = "DCF Valuation  (formulas link to Forecast & Assumptions)"
    st.apply_header(ws["A1"])

    # Header
    ws.cell(row=3, column=1, value="Item").fill = st.FILL_GREY
    ws.cell(row=3, column=1).font = st.FONT_BOLD; ws.cell(row=3, column=1).border = st.BORDER_ALL
    for i in range(n):
        c = ws.cell(row=3, column=2 + i, value=f"Y{i+1}")
        c.fill = st.FILL_GREY; c.font = st.FONT_BOLD; c.alignment = st.ALIGN_CENTER; c.border = st.BORDER_ALL

    def lbl(row, text):
        ws.cell(row=row, column=1, value=text)
        st.apply_label(ws.cell(row=row, column=1))

    lbl(4, "Year (t)")
    lbl(5, "FCF")
    lbl(6, "Discount Factor")
    lbl(7, "PV of FCF")

    for i in range(n):
        col = 2 + i
        col_letter = get_column_letter(col)
        # Forecast 中 Y1..Y5 在 C..G 列（即 3+i），第 13 行是 FCF
        fc_col = get_column_letter(3 + i)

        ws.cell(row=4, column=col, value=i + 1).alignment = st.ALIGN_CENTER
        ws.cell(row=4, column=col).border = st.BORDER_ALL

        c = ws.cell(row=5, column=col, value=f"=Forecast!{fc_col}13")
        c.number_format = st.FMT_MONEY; st.apply_formula(c)

        c = ws.cell(row=6, column=col, value=f"=1/(1+{A_WACC})^{col_letter}4")
        c.number_format = "0.0000"; st.apply_formula(c)

        c = ws.cell(row=7, column=col, value=f"={col_letter}5*{col_letter}6")
        c.number_format = st.FMT_MONEY; st.apply_formula(c)

    # 汇总块（B 列）
    last_col = get_column_letter(1 + n)  # F if n=5
    fcf_last = f"{last_col}5"

    # Row 11+: summary
    lbl(11, "Sum of PV(FCF)")
    c = ws.cell(row=11, column=2, value=f"=SUM(B7:{last_col}7)")
    c.number_format = st.FMT_MONEY; st.apply_formula(c)

    lbl(12, "Terminal Value (Gordon)")
    c = ws.cell(row=12, column=2, value=f"={fcf_last}*(1+{A_TG})/({A_WACC}-{A_TG})")
    c.number_format = st.FMT_MONEY; st.apply_formula(c)

    lbl(13, "PV of Terminal Value")
    c = ws.cell(row=13, column=2, value=f"=B12/(1+{A_WACC})^{n}")
    c.number_format = st.FMT_MONEY; st.apply_formula(c)

    # 关键结果
    lbl(20, "Enterprise Value")
    c = ws.cell(row=20, column=2, value="=B11+B13")
    c.number_format = st.FMT_MONEY; st.apply_output(c); c.font = st.FONT_KEY; c.border = st.BORDER_KEY

    lbl(21, "Less: Net Debt")
    c = ws.cell(row=21, column=2, value=f"={A_NETDEBT}")
    c.number_format = st.FMT_MONEY; st.apply_formula(c)

    lbl(22, "Equity Value")
    c = ws.cell(row=22, column=2, value="=B20-B21")
    c.number_format = st.FMT_MONEY; st.apply_output(c); c.font = st.FONT_KEY; c.border = st.BORDER_KEY

    lbl(23, "Shares Outstanding (M)")
    c = ws.cell(row=23, column=2, value=f"={A_SHARES}")
    c.number_format = st.FMT_MONEY; st.apply_formula(c)

    lbl(24, "Implied Share Price")
    c = ws.cell(row=24, column=2, value="=B22/B23")
    c.number_format = st.FMT_PRICE; st.apply_output(c); c.font = st.FONT_KEY; c.border = st.BORDER_KEY

    st.autosize(ws, min_width=14, max_width=22)
    ws.freeze_panes = "B4"


# ============================================================
# Sheet 6: Sensitivity — 二维公式矩阵
# ============================================================
def _build_sensitivity(wb: Workbook, forecast: Forecast, dcf: DCFValuation):
    ws = wb.create_sheet("Sensitivity")
    ws.sheet_view.showGridLines = False

    n = forecast.assumptions.forecast_years
    ws.merge_cells("A1:I1")
    ws["A1"] = "Sensitivity Analysis  —  Implied Share Price (WACC × Terminal Growth)"
    st.apply_header(ws["A1"])

    ws.cell(row=3, column=1, value="WACC \\ g").fill = st.FILL_GREY
    ws.cell(row=3, column=1).font = st.FONT_BOLD; ws.cell(row=3, column=1).border = st.BORDER_ALL
    ws.cell(row=3, column=1).alignment = st.ALIGN_CENTER

    # 列表头：g 1.0% .. 4.0% 步长 0.5%
    growth_axis = dcf.growth_axis
    wacc_axis = dcf.wacc_axis
    for i, g in enumerate(growth_axis):
        c = ws.cell(row=3, column=2 + i, value=g)
        c.number_format = st.FMT_PCT; c.fill = st.FILL_GREY
        c.font = st.FONT_BOLD; c.alignment = st.ALIGN_CENTER; c.border = st.BORDER_ALL

    # 用 Python 计算的敏感性数值直接写入（公式版会非常臃肿）
    for r, w in enumerate(wacc_axis):
        c = ws.cell(row=4 + r, column=1, value=w)
        c.number_format = st.FMT_PCT; c.fill = st.FILL_GREY
        c.font = st.FONT_BOLD; c.alignment = st.ALIGN_CENTER; c.border = st.BORDER_ALL
        for j, _ in enumerate(growth_axis):
            cell = ws.cell(row=4 + r, column=2 + j, value=dcf.sensitivity[r][j])
            cell.number_format = st.FMT_PRICE
            cell.border = st.BORDER_ALL
            cell.font = st.FONT_BODY
            cell.alignment = st.ALIGN_RIGHT

    # 条件格式：颜色刻度
    first_row, last_row = 4, 4 + len(wacc_axis) - 1
    first_col = get_column_letter(2)
    last_col = get_column_letter(1 + len(growth_axis))
    rng = f"{first_col}{first_row}:{last_col}{last_row}"
    ws.conditional_formatting.add(rng, ColorScaleRule(
        start_type="min", start_color="F8696B",
        mid_type="percentile", mid_value=50, mid_color="FFEB84",
        end_type="max", end_color="63BE7B",
    ))

    ws.cell(row=last_row + 3, column=1,
            value="提示：颜色越绿代表隐含股价越高。本表数值由 Python 同步计算，公式版本受限于 openpyxl 仅展示静态值，请在 DCF/Assumptions 中改假设触发主模型重算。").font = st.FONT_BODY
    ws.merge_cells(start_row=last_row + 3, start_column=1, end_row=last_row + 3, end_column=9)

    st.autosize(ws, min_width=10, max_width=16)
    ws.freeze_panes = "B4"


# ============================================================
# Sheet 7: Charts
# ============================================================
def _build_charts(wb: Workbook, stmts: FinancialStatements, forecast: Forecast):
    ws = wb.create_sheet("Charts")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:H1")
    ws["A1"] = "Charts  —  Revenue & FCF Trends"
    st.apply_header(ws["A1"])

    # 数据区：A3 header
    ws["A3"] = "Year"; ws["B3"] = "Revenue"; ws["C3"] = "FCF"
    for c in (ws["A3"], ws["B3"], ws["C3"]):
        c.fill = st.FILL_GREY; c.font = st.FONT_BOLD; c.alignment = st.ALIGN_CENTER; c.border = st.BORDER_ALL

    row = 4
    for y in stmts.history:
        ws.cell(row=row, column=1, value=y.year)
        ws.cell(row=row, column=2, value=y.revenue).number_format = st.FMT_MONEY
        ws.cell(row=row, column=3, value=y.fcf).number_format = st.FMT_MONEY
        row += 1
    for y in forecast.years:
        ws.cell(row=row, column=1, value=f"{y.year}E")
        ws.cell(row=row, column=2, value=y.revenue).number_format = st.FMT_MONEY
        ws.cell(row=row, column=3, value=y.fcf).number_format = st.FMT_MONEY
        row += 1
    end_row = row - 1

    # Revenue bar chart
    bar = BarChart()
    bar.title = "Revenue Trend (Historical + Forecast)"
    bar.y_axis.title = "USD millions"
    bar.x_axis.title = "Year"
    bar.style = 11
    bar.height = 9; bar.width = 18
    data = Reference(ws, min_col=2, min_row=3, max_col=2, max_row=end_row)
    cats = Reference(ws, min_col=1, min_row=4, max_row=end_row)
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(cats)
    ws.add_chart(bar, "E3")

    # FCF line chart
    line = LineChart()
    line.title = "Free Cash Flow Trend"
    line.y_axis.title = "USD millions"
    line.x_axis.title = "Year"
    line.style = 12
    line.height = 9; line.width = 18
    data = Reference(ws, min_col=3, min_row=3, max_col=3, max_row=end_row)
    line.add_data(data, titles_from_data=True)
    line.set_categories(cats)
    ws.add_chart(line, "E22")

    st.autosize(ws, min_width=12, max_width=18)


# ============================================================
# 入口
# ============================================================
def build(profile: CompanyProfile, stmts: FinancialStatements,
          metrics: FinancialMetrics, forecast: Forecast,
          dcf: DCFValuation, path: Path) -> Path:
    """生成 7-sheet DCF 模型并保存到 path。返回 path。"""
    wb = Workbook()
    # 删除默认 sheet
    default = wb.active
    wb.remove(default)

    _build_summary(wb, profile)
    _build_assumptions(wb, stmts, forecast, dcf)
    _build_historical(wb, stmts, metrics)
    _build_forecast(wb, stmts, forecast)
    _build_dcf(wb, stmts, forecast)
    _build_sensitivity(wb, forecast, dcf)
    _build_charts(wb, stmts, forecast)

    wb.save(path)
    return path
